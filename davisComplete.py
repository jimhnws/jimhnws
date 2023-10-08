#!/usr/bin/env python
# coding: utf-8

# In[20]:


import calcOneDay
import getDays
from datetime import datetime, timedelta

# Calculate the time and date for end of day calculations

xy = calcOneDay.calcOneDay()
start, end = (xy[0], xy[1])

todayInfo = getDays.getToday()
yesterdayInfo = getDays.getYesterday()
print(todayInfo)
print(yesterdayInfo)
tomorrowInfo = getDays.getTomorrow()

month, month_num, date, year = todayInfo[0], todayInfo[1], todayInfo[2], todayInfo[3]
yesterday = yesterdayInfo[2]
yesterday = int(yesterday)
nextDay = tomorrowInfo[2]
nextDay = int(nextDay)


# In[21]:


import collections
import hashlib
import hmac
import time
from datetime import datetime
import requests
import json
import dataFile

parameters = {
  "api-key": "vy8jbrjsxlbwgojepq3vfyfqfywyhvbd", 
  "api-secret": "sdqfm6wdfy9w0pqp2vdka38o6b4vcsvc",
  "station-id": 81211, 
  "end-timestamp": end,
  "start-timestamp": start,
  "t": int(time.time())
}

parameters = collections.OrderedDict(sorted(parameters.items()))

for key in parameters:
    print("Parameter name: \"{}\" has value \"{}\"".format(key, parameters[key]))

apiSecret = parameters["api-secret"];
parameters.pop("api-secret", None);

data = ""
for key in parameters:
    data = data + key + str(parameters[key])

print("Data string to hash is: \"{}\"".format(data))
print('\n')

"""
Calculate the HMAC SHA-256 hash that will be used as the API Signature.
"""
apiSignature = hmac.new(
  apiSecret.encode('utf-8'),
  data.encode('utf-8'),
  hashlib.sha256
).hexdigest()

"""
Let's see what the final API Signature looks like.
"""
print("API Signature is: \"{}\"".format(apiSignature))
print('\n')

# Building the URL to get the station

first_part = ('https://api.weatherlink.com/v2/historic/81211?')
api_key = ('api-key=vy8jbrjsxlbwgojepq3vfyfqfywyhvbd')
add_apisig = ('&api-signature=')
add_t = ('&t='+ str(int(time.time())))

start1 = "&start-timestamp=" + start
end1 = "&end-timestamp=" + end

#
URLfinal = (first_part + api_key + add_t + start1 + end1 + add_apisig + apiSignature)
# print(URLfinal)

r =  requests.get(URLfinal)
davisAPI = (r.json())


# In[11]:


import numpy as np
import pandas as pd
from pandas import DataFrame, Series
import matplotlib.pyplot as plt
import sqlalchemy
from dateutil.tz import tzutc, tzlocal
import pytz

a = davisAPI['sensors']    
b = a[1]
c = (b['data'])

df = pd.DataFrame(c) 
df.rename(columns = {'ts':'timestamp'}, inplace = True)

timezone = pytz.timezone("America/New_York")
df['timeGroup'] = pd.to_datetime(df['timestamp'], unit='s')
df['timeGroup'] = df['timeGroup'].dt.tz_localize('UTC').dt.tz_convert('US/Eastern')
df['localTime'] = df['timeGroup'].dt.strftime('%I:%M %p')

df = df.loc[:,['timestamp', 'temp_hi', 'temp_hi_at','temp_lo', 'temp_lo_at', 'rainfall_in', 'dew_point_hi', 'dew_point_lo',  'rain_rate_hi_in', 'rain_rate_hi_at', 'timeGroup', 'localTime']]


# In[22]:


import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from pandas import DataFrame, Series
import math

max_temp  = (df.sort_values(by='temp_hi', ascending=False))
max_T = max_temp.iloc[:1]
max_T_time = int(max_T['temp_hi_at'])
maxT = max_T['temp_hi'].values[0]
maxT = round(maxT)

min_temp  = (df.sort_values(by='temp_lo', ascending=True))
min_T = min_temp.iloc[:1]
min_T_time = int(min_T['temp_lo_at'])
minT = min_T['temp_lo'].values[0]
minT = round(minT)

dew_max = (df.sort_values(by='dew_point_hi', ascending=False))
dew_max1 = dew_max.iloc[:1]
dewMax = dew_max1['dew_point_hi'].values[0]
dewMaxT = round(dewMax)

dew_min = (df.sort_values(by='dew_point_lo', ascending=True))
dew_min1 = dew_min.iloc[:1]
dewMin = dew_min1['dew_point_lo'].values[0]
dewMinT = round(dewMin)

rain = df['rainfall_in'].sum()
avgTemp = math.ceil((int(maxT + minT)/2))

hdd = (65 - avgTemp)
if hdd < 0:
    hdd = 0
cdd = (avgTemp - 65)
if cdd < 0:
    cdd = 0  


# In[ ]:


import sqlalchemy
import mysql.connector
import sqlite3

df2 = pd.DataFrame(columns = ['Year', 'Month', 'Date', 'High', 'Low', 'Rainfall', 'Max_Dew_Point'])
newRow = pd.DataFrame({'Year': year, 'Month': month_num, 'Date': yesterday, 'High': maxT, 'Low': minT, 'Rainfall' : rain, 'Max_Dew_Point': dewMaxT }, index = [yesterday])
df2 = pd.concat([newRow, df2]).reset_index(drop = True)
print(df2)

database_username = 'chuckwx'
database_password = 'jfr716!!00'
database_ip       = '3.135.162.69'
database_name     = 'davisf6'
database_connection = sqlalchemy.create_engine('mysql+mysqlconnector://{0}:{1}@{2}/{3}'.
                                               format(database_username, database_password, 
                                                      database_ip, database_name), connect_args={'connect_timeout': 30})
df2.to_sql(con=database_connection, name='davisTest', if_exists='append', index = False)


# In[27]:


import openpyxl
from openpyxl import load_workbook
import excelFilename
import calcTimeNow
import getNameNumbers

#
# Create the month name for the xlsx filename
#

#gg = getNameNumbers.sqlWrite()
#xls_filename, xls_fullfile, path_name = gg[0], gg[1], gg[2]
#print(xls_filename, xls_fullfile, path_name)

xls_filename = 'davisTest.xlsx'
path_name = '/home/ec2-user/'
xls_fullfile = f'{path_name}{xls_filename}'
print(xls_fullfile)

wb = openpyxl.load_workbook(xls_fullfile)
sheet = wb.active

# Write headers first...
a1 = sheet['A1']
a1.value = "Year"
b1 = sheet['B1']
b1.value = year
c1 = sheet['C1']
c1.value = 'Month'
d1 = sheet['D1']
d1.value = month

a3 = sheet['A3']
a3.value = "Date"
b3 = sheet['B3']
b3.value = 'High'
c3 = sheet['C3']
c3.value = 'Low'
d3 = sheet['D3']
d3.value = 'Average'

e3 = sheet['E3']
e3.value = "HDD"
f3 = sheet['F3']
f3.value = 'CDD'
g3 = sheet['G3']
g3.value = 'Rainfall'
h3 = sheet['H3']
h3.value = 'Max_Dew_Pt'
i3 = sheet['I3']
i3.value = 'Min_Dew_Pt'

k4 = sheet['K4']
k4.value = "Highs >=90"
k5 = sheet['K5']
k5.value = "Highs <= 32"
k6 = sheet['K6']
k6.value = 'Lows <= 32'
k7 = sheet['K7']
k7.value = 'Lows <= 0'

k13 = sheet['K14']
k13.value = "Total Rainfall"
k14 = sheet['K15']
k14.value = "rain>=0.01"
k15 = sheet['K16']
k15.value = 'rain>=0.10'
k16 = sheet['K17']
k16.value = 'rain>=0.50'
k17= sheet['K18']
k17.value = 'rain>=1.00'
k23 = sheet['K24']
k23.value = 'Monthly Average'
k24 = sheet['K25']
k24.value = 'Departure'

m3 = sheet['M4']
m3.value = "High"
m4 = sheet['M5']
m4.value = "Low"
m13 = sheet['M14']
m13.value = "Max Rain"
m23 = sheet['M24']
m23.value = "Monthy Rainfall"
m24 = sheet['M25']
m24.value = "Departure"

o3 = sheet['O4']
o3.value = "Date"
o4 = sheet['O5']
o4.value = "Date"


# Calculate the date and write the data...
offset_day = (int(date) + 2)

date1 = sheet.cell(row = offset_day, column = 1)
date1.value = yesterday
maxTT = sheet.cell(row = offset_day, column = 2)
maxTT.value = maxT
minTT = sheet.cell(row = offset_day, column = 3)
minTT.value = minT
avgT = sheet.cell(row = offset_day, column = 4)
avgT.value = avgTemp
hdd1 = sheet.cell(row = offset_day, column = 5)
hdd1.value = hdd
cdd1 = sheet.cell(row = offset_day, column = 6)
cdd1.value = cdd
totRR = sheet.cell(row = offset_day, column = 7)
totRR.value = rain
dewMaxTT = sheet.cell(row = offset_day, column = 8)
dewMaxTT.value = dewMaxT
dewMinTT = sheet.cell(row = offset_day, column = 9)
dewMinTT.value = dewMinT

wb.save(xls_fullfile)


# In[ ]:


import setUpHTML

# Read the Excel file as a possible pandas dataframe and html file

html_path = '/var/www/html/000/'

df1 = pd.read_excel(xls_fullfile, skiprows = 2, names = ['Date','High','Low','Average','HDD','CDD','Rainfall','Max Dew Pt','Min Dew Pt','dead2','dead3','dead4','dead5','dead6','dead7'])
df1 = df1.drop(df1.columns[[9,10,11,12,13,14]], axis = 1)
df1
df1.to_html(f'{html_path}throttle.html', index = False) 


# In[7]:


import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from pandas import DataFrame, Series
import pymysql as dbapi
import sys
import csv
from tabulate import tabulate

#
# Get normal highs and lows
#

QUERY = """SELECT * FROM avgHiLo 
           WHERE Month = %s 
           AND Day = %s""" % (month_num, date)

print(QUERY)


db = dbapi.connect(host='3.135.162.69',user='chuckwx',passwd='jfr716!!00', database = 'trweather')

cur = db.cursor()
cur.execute(QUERY)
result = cur.fetchall()
print(result)

dataset = result[0]
nmlHi = int(dataset[3])
nmlLo = int(dataset[4])

#
# Get the record high for the date
#

QUERY1 = """SELECT * FROM recHigh 
           WHERE Month = %s 
           AND Day = %s""" % (month_num, date)


db = dbapi.connect(host='3.135.162.69',user='chuckwx',passwd='jfr716!!00', database = 'trweather')

cur = db.cursor()
cur.execute(QUERY1)
result1 = cur.fetchall()
recordHigh = result1[0]
recHigh = int(recordHigh[1])
recHighYear = int(recordHigh[4])

#
# Get the record low for the date
#

QUERY2 = """SELECT * FROM recLow 
           WHERE Month = %s 
           AND Day = %s""" % (month_num, date)


db = dbapi.connect(host='3.135.162.69',user='chuckwx',passwd='jfr716!!00', database = 'trweather')

cur = db.cursor()
cur.execute(QUERY2)
result2 = cur.fetchall()
recYearNum =  len(result2)
recordLow = result2[0]
recLow = int(recordLow[1])
recLowYear = int(recordLow[4])

#
# Get the record rainfall for the date
#

QUERY3 = """SELECT * FROM recRain 
           WHERE Month = %s 
           AND Day = %s""" % (month_num, date)


db = dbapi.connect(host='3.135.162.69',user='chuckwx',passwd='jfr716!!00', database = 'trweather')

cur = db.cursor()
cur.execute(QUERY3)
result3 = cur.fetchall()
#sandbox1.recordRain(result3)
recordRain = result3[0]
recRain = recordRain[1]
recRainYear = int(recordRain[4])


# In[ ]:


import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from pandas import DataFrame, Series
import pymysql as dbapi
import sys
import csv
from tabulate import tabulate
import sandbox1
import sandbox2

'''
climo_data = [["Month", "Day", "Year", "High", "Low", "Avg", "HDD", "CDD", "Rain"],
             [month, date, year, maxT, minT, avgTemp, hdd, cdd, corR]]

with open('/var/www/html/000/climo.html', 'w') as f:
    f.write(tabulate(climo_data, headers = 'firstrow', tablefmt = 'html'))
    

record_data = [["Month", "Day", "Year", "Record High", "Year", "Record Low", "Year", "Record Rainfall", "Year"],
             [month, nextDay, year, recHigh, recHighYear, recLow, recLowYear, recRain, recRainYear]]
            


with open('/var/www/html/000/day_records.html', 'w') as f:
    f.write(tabulate(record_data, headers = 'firstrow', tablefmt = 'html'))'

with open('/var/www/html/000/climoTest1.html','w') as outfile1: 
    print(f'This is the daily almanac for {month} {date}, {year}', file = outfile1)
    print('\n', file = outfile1)
    print(f'The high today was {maxT} at {hiTime}', file = outfile1)
    print(f'The low today was {minT} at {loTime}', file = outfile1)
    print(f'The average temperature was {avgTemp}', file = outfile1)
    print(f'The rainfall today was {corR} inches', file = outfile1)
    print(f'There were {hdd} heating degree days today', file = outfile1)
    print(f'There were {cdd} cooling degree days today', file = outfile1)
'''

nmlData = sandbox2.sandbox2()
print("This is the value of nmlData: ", nmlData)
nmlHi = nmlData[3]
nmlLo = nmlData[4]

highData = sandbox1.recordHigh()
lowData = sandbox1.recordLow()
rainData = sandbox1.recordRain()

highPhrase = highData[2]
lowPhrase = lowData[2]
rainPhrase = rainData[2]

with open('/var/www/html/000/climoDavisText.txt','w') as outfile1: 
    print(f'Daily almanac for {month} {yesterday}, {year}', file = outfile1)
    print('\n', file = outfile1)
    print(f'The high yesterday was {maxT} degrees', file = outfile1)
    print(f'The low yesterday was {minT} degrees', file = outfile1)
    print(f'The average temperature was {avgTemp} degrees', file = outfile1)
    print(f'The rainfall yesterday was {("%.2f" % rain)} inches', file = outfile1)
    print(f'There were {hdd} heating degree days', file = outfile1)
    print(f'There were {cdd} cooling degree days', file = outfile1)
    print('\n', file = outfile1)
    
    print(f'Normal and Record information for {month} {date}, {year}', file = outfile1)
    print('\n', file = outfile1)
    print(f'The normal high for today is {nmlHi} degrees', file = outfile1)
    print(f'The normal low for today is {nmlLo} degrees' , file = outfile1)
    print('\n', file = outfile1)
    print(highPhrase, file = outfile1)
    print(lowPhrase, file = outfile1)
    print(rainPhrase, file = outfile1)  

